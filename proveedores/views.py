# views.py
from django.shortcuts import render, redirect, get_object_or_404
from .models import Proveedor
from .forms import ProveedorForm
from django.db.models import Q
from django.core.paginator import Paginator
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache
from login.decorators import role_required
from django.db import transaction
import random
import string
from django.contrib import messages

@never_cache
@login_required
@role_required(['ADMIN'])
def exportar_proveedores_excel(request):
    query = request.GET.get('q', '').strip()
    proveedores_qs = Proveedor.objects.all()

    if query:
        proveedores_qs = proveedores_qs.filter(
            Q(razon_social__icontains=query) |
            Q(rut__icontains=query)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proveedores"

    headers = ["RUT", "Razón Social", "Email", "Teléfono", "Ciudad"]
    ws.append(headers)

    for p in proveedores_qs.iterator():
        ws.append([p.rut, p.razon_social, p.email, p.telefono, p.ciudad])

    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 2)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=proveedores.xlsx'
    wb.save(response)
    return response

@never_cache
@login_required
@role_required(['ADMIN'])
def lista_proveedores(request):
    query = request.GET.get('q', '').strip()
    proveedores_qs = Proveedor.objects.all()

    if query:
        proveedores_qs = proveedores_qs.filter(
            Q(razon_social__icontains=query) |
            Q(rut__icontains=query)
        )

    paginator = Paginator(proveedores_qs, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {"page_obj": page_obj, "query": query}

    if request.headers.get("HX-Request") == "true":
        return render(request, "proveedores/_grid.html", context)

    return render(request, "proveedores/lista.html", context)


# views.py
@never_cache
@login_required
@role_required(['ADMIN'])
def crear_proveedor(request):
    if request.method == 'POST':
        print("Formulario recibido")
        form = ProveedorForm(request.POST)
        if form.is_valid():
            print("Formulario válido")
            form.save()
            return redirect('lista_proveedores')
        else:
            print("Errores del formulario:", form.errors)
    else:
        form = ProveedorForm()
    return render(request, 'proveedores/formulario.html', {'form': form})

@never_cache
@login_required
@role_required(['ADMIN'])
def editar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            return redirect('lista_proveedores')
    else:
        form = ProveedorForm(instance=proveedor)
    return render(request, 'proveedores/formulario.html', {'form': form})

@never_cache
@login_required
@role_required(['ADMIN'])
def eliminar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == 'POST':
        proveedor.delete()
        return redirect('lista_proveedores')
    return render(request, 'proveedores/eliminar_proveedor.html', {'proveedor': proveedor})

@never_cache
@login_required
@role_required(['ADMIN'])
def test_form(request):
    if request.method == 'POST':
        print("Formulario recibido")  # Esto debe aparecer en consola
    return render(request, 'proveedores/test_form.html')


@never_cache
@login_required
@role_required(['ADMIN'])
def bulk_create_proveedores(request):
    """Create N test providers for stress testing. POST only."""
    if request.method != 'POST':
        return redirect('lista_proveedores')

    try:
        count = int(request.POST.get('count', 1000))
    except (TypeError, ValueError):
        count = 1000
    # Safety cap
    count = max(1, min(count, 20000))
    objs = []
    base_rut = 10000000
    for i in range(count):
        rut = str(base_rut + i)
        razon_social = f"Proveedor de Prueba {rut}"
        nombre_fantasia = razon_social
        email = f"prov{rut}@example.com"
        telefono = str(900000000 + (i % 10000000))
        direccion = f"Calle Prueba {i}"
        ciudad = "Santiago"
        pais = "Chile"
        plazos_pago = "30 días"
        moneda = "CLP"
        objs.append(Proveedor(
            rut=rut,
            razon_social=razon_social,
            nombre_fantasia=nombre_fantasia,
            email=email,
            telefono=telefono,
            direccion=direccion,
            ciudad=ciudad,
            pais=pais,
            plazos_pago=plazos_pago,
            moneda=moneda,
        ))

    with transaction.atomic():
        Proveedor.objects.bulk_create(objs, batch_size=500)
    messages.success(request, f'Se crearon {count} proveedores de prueba.')
    return redirect('lista_proveedores')


@never_cache
@login_required
@role_required(['ADMIN'])
def bulk_delete_proveedores(request):
    """Delete providers created by the stress loader (heuristic: razon_social startswith 'Proveedor de Prueba' or email ending with @example.com)."""
    if request.method != 'POST':
        return redirect('lista_proveedores')

    qs = Proveedor.objects.filter(Q(razon_social__startswith='Proveedor de Prueba') | Q(email__endswith='@example.com'))
    deleted_count, _ = qs.delete()
    messages.success(request, f'Se eliminaron {deleted_count} proveedores de prueba.')
    return redirect('lista_proveedores')