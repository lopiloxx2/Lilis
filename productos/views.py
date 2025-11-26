from django.shortcuts import render, get_object_or_404, redirect
from .models import Categoria, Producto, Usuario, Venta, DetalleVenta
from .forms import ProductoForm
from .forms import VentaForm
from django.db.models import Q, Sum
from django.db import transaction
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.core.mail import send_mail
from django.conf import settings
import logging
import smtplib
import openpyxl
from openpyxl.utils import get_column_letter
from .forms import UsuarioForm 
from django.contrib.auth.hashers import make_password, check_password
from django.contrib.auth.models import User as AuthUser
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache
from login.decorators import role_required
from django.shortcuts import get_object_or_404
import secrets
import string
import random
from Lilis.auth_utils import generate_temporary_password
from django.urls import reverse
from decimal import Decimal
from .models import InventoryMovement, MovementItem, Bodega, Lote
from .models import Producto
from proveedores.models import Proveedor
from django.contrib import messages
from types import SimpleNamespace
from django.utils import timezone
from datetime import timedelta
from django.views.decorators.cache import never_cache

@never_cache
@login_required
def inicio(request):
    return render(request, 'productos/index.html')


@never_cache
@login_required
@role_required(['ADMIN','VENDEDOR','CAJA'])
def inventory_list(request):
    # Handle creation of a simple movement (single item) from the form
    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        referencia = request.POST.get('referencia', '')
        notas = request.POST.get('notas', '')
        proveedor_id = request.POST.get('proveedor')
        bodega_id = request.POST.get('bodega')
        producto_id = request.POST.get('producto')
        lote_code = request.POST.get('lote')
        cantidad = request.POST.get('cantidad')
        precio = request.POST.get('precio_unitario') or '0'

        try:
            proveedor = Proveedor.objects.get(pk=proveedor_id) if proveedor_id else None
        except Proveedor.DoesNotExist:
            proveedor = None

        try:
            bodega = Bodega.objects.get(pk=bodega_id) if bodega_id else None
        except Bodega.DoesNotExist:
            bodega = None

        try:
            producto = Producto.objects.get(pk=producto_id) if producto_id else None
        except Producto.DoesNotExist:
            producto = None

        lote_obj = None
        if lote_code and producto:
            lote_obj, created = Lote.objects.get_or_create(producto=producto, codigo_lote=lote_code, defaults={'cantidad': 0, 'proveedor': proveedor})

        try:
            qty = int(cantidad or 0)
        except ValueError:
            qty = 0

        try:
            price = float(precio)
        except ValueError:
            price = 0

        # Validate and create movement atomically to avoid partial updates
        if not producto or qty == 0:
            messages.error(request, 'Seleccione un producto válido y cantidad mayor a 0.')
            return redirect('inventory_list')

        # For SALIDA and TRANSFERENCIA ensure enough stock/lote
        if tipo in ['SALIDA', 'TRANSFERENCIA']:
            if (producto.stock or 0) < qty:
                messages.error(request, f'Stock insuficiente para {producto.nombre} (disponible: {producto.stock}).')
                return redirect('inventory_list')
            if lote_obj and (lote_obj.cantidad or 0) < qty:
                messages.error(request, f'Cantidad insuficiente en lote {lote_obj.codigo_lote} (disponible: {lote_obj.cantidad}).')
                return redirect('inventory_list')

        with transaction.atomic():
            movimiento = InventoryMovement.objects.create(
                tipo=tipo or 'AJUSTE',
                fecha=timezone.now(),
                usuario=request.user.username if request.user.is_authenticated else '',
                proveedor=proveedor,
                bodega_origen=bodega if tipo in ['SALIDA','TRANSFERENCIA'] else None,
                bodega_destino=bodega if tipo in ['INGRESO','TRANSFERENCIA'] else None,
                referencia=referencia,
                notas=notas,
            )

            MovementItem.objects.create(
                movimiento=movimiento,
                producto=producto,
                lote=lote_obj,
                cantidad=qty,
                precio_unitario=price,
            )

            # Adjust stock and lote quantities (INGRESO increments, SALIDA decrements)
            if movimiento.tipo == 'INGRESO':
                producto.stock = (producto.stock or 0) + qty
                producto.save()
                if lote_obj:
                    lote_obj.cantidad = (lote_obj.cantidad or 0) + qty
                    lote_obj.save()
            elif movimiento.tipo == 'SALIDA':
                producto.stock = (producto.stock or 0) - qty
                producto.save()
                if lote_obj:
                    lote_obj.cantidad = (lote_obj.cantidad or 0) - qty
                    lote_obj.save()

        messages.success(request, 'Movimiento creado correctamente.')
        return redirect('inventory_list')

    # list recent movements
    movimientos = InventoryMovement.objects.select_related('proveedor','bodega_origen','bodega_destino').order_by('-fecha')[:200]

    # statistics
    # Compute today's range in local time to avoid date mismatches with DB timezone
    now_local = timezone.localtime(timezone.now())
    start_of_day = now_local.replace(hour=0, minute=0, second=0, microsecond=0)
    end_of_day = start_of_day + timedelta(days=1)
    movimientos_hoy = InventoryMovement.objects.filter(fecha__gte=start_of_day, fecha__lt=end_of_day).count()
    stock_total_demo = Producto.objects.aggregate(total=Sum('stock'))['total'] or 0
    productos_unicos = Producto.objects.count()

    proveedores = Proveedor.objects.all()[:200]
    # Ensure default bodegas exist
    default_bodegas = [
        "Lili's Bebidas",
        "Lili's Chocolatería",
        "Lili's Pastelería",
    ]
    for name in default_bodegas:
        Bodega.objects.get_or_create(nombre=name)

    bodegas = Bodega.objects.all()[:50]
    productos = Producto.objects.all()[:200]

    context = {
        'movimientos': movimientos,
        'movimientos_hoy': movimientos_hoy,
        'stock_total_demo': stock_total_demo,
        'productos_unicos': productos_unicos,
        'now': now_local,
        'proveedores': proveedores,
        'bodegas': bodegas,
        'productos': productos,
    }
    return render(request, 'productos/inventory_list.html', context)


@never_cache
@login_required
@role_required(['ADMIN','VENDEDOR','CAJA'])
def inventory_detail(request, pk):
    movimiento = get_object_or_404(InventoryMovement, pk=pk)
    items = movimiento.items.select_related('producto','lote').all()
    return render(request, 'productos/inventory_detail.html', {'movimiento': movimiento, 'items': items})

@never_cache
@login_required
@role_required(['ADMIN'])
def editar_usuario(request, pk):
    usuario = get_object_or_404(Usuario, pk=pk)
    if request.method == 'POST':
        form = UsuarioForm(request.POST, instance=usuario)
        if form.is_valid():
            u = form.save()
            # sincronizar con auth.User si existe
            try:
                auth_user = AuthUser.objects.get(username=u.username)
                auth_user.email = u.email
                auth_user.save()
            except AuthUser.DoesNotExist:
                pass
            return redirect('lista_usuarios')
    else:
        form = UsuarioForm(instance=usuario)
    return render(request, 'usuarios/editar_usuario.html', {'form': form, 'usuario': usuario})

@never_cache
@login_required
@role_required(['ADMIN'])
def eliminar_usuario(request, pk):
    usuario = get_object_or_404(Usuario, pk=pk)
    if request.method == 'POST':
        # Prevent deleting the currently logged-in user to avoid accidental logout
        if request.user.is_authenticated and usuario.username == request.user.username:
            messages.error(request, 'No puede eliminar su propio usuario desde esta pantalla.')
            return redirect('lista_usuarios')

        # eliminar usuario del modelo de la app
        usuario.delete()
        # eliminar usuario de auth si existe (pero no el que está logueado)
        try:
            auth_user = AuthUser.objects.get(username=usuario.username)
            if request.user.is_authenticated and auth_user.username == request.user.username:
                # don't delete the currently logged-in auth user
                pass
            else:
                auth_user.delete()
        except AuthUser.DoesNotExist:
            pass
        messages.success(request, 'Usuario eliminado correctamente.')
        return redirect('lista_usuarios')
    return render(request, 'usuarios/eliminar_usuario.html', {'usuario': usuario})

def exportar_productos_excel(request):
    query = request.GET.get('q', '').strip()
    productos_qs = Producto.objects.select_related("categoria").all()

    if query:
        productos_qs = productos_qs.filter(
            Q(nombre__icontains=query) |
            Q(sku__icontains=query) |
            Q(marca__icontains=query)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    headers = ["SKU", "Nombre", "Categoría", "Precio", "Stock", "Marca"]
    ws.append(headers)

    for p in productos_qs.iterator():
        ws.append([
            p.sku,
            p.nombre,
            p.categoria.nombre_categoria if p.categoria else "",
            float(p.precio_venta) if p.precio_venta is not None else 0.0,
            p.stock if p.stock is not None else 0,
            p.marca or ""
        ])

    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 2)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=productos.xlsx'
    wb.save(response)
    return response

def ver_productos(request):
    productos = Producto.objects.all()
    print("=== PRODUCTOS ENCONTRADOS ===")
    print(productos)
    for p in productos:
        print(p.nombre)
    return render(request, 'productos/productos.html', {'productos': productos})


@never_cache
@login_required
@role_required(['ADMIN','VENDEDOR','CAJA'])
def lista_productos(request):
    query = request.GET.get('q', '').strip()
    productos_qs = Producto.objects.select_related("categoria").all()

    if query:
        productos_qs = productos_qs.filter(
            Q(nombre__icontains=query) |
            Q(sku__icontains=query) |
            Q(marca__icontains=query)
        )

    paginator = Paginator(productos_qs, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {"page_obj": page_obj, "query": query}

    if request.headers.get("HX-Request") == "true":
        return render(request, "productos/_grid.html", context)

    return render(request, "productos/productos.html", context)


@never_cache
@login_required
@role_required(['ADMIN'])
def lista_usuarios(request):
    query = request.GET.get('q', '').strip()
    usuarios_qs = Usuario.objects.all()

    if query:
        usuarios_qs = usuarios_qs.filter(
            Q(username__icontains=query) |
            Q(email__icontains=query) |
            Q(nombres__icontains=query) |
            Q(apellidos__icontains=query) |
            Q(telefono__icontains=query)
        )

    paginator = Paginator(usuarios_qs, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {"page_obj": page_obj, "query": query}

    if request.headers.get("HX-Request") == "true":
        return render(request, "usuarios/_grid_usuarios.html", context)

    return render(request, "usuarios/usuarios.html", context)

def exportar_usuarios_excel(request):
    query = request.GET.get('q', '').strip()
    usuarios_qs = Usuario.objects.all()

    if query:
        usuarios_qs = usuarios_qs.filter(
            Q(username__icontains=query) |
            Q(email__icontains=query) |
            Q(nombres__icontains=query) |
            Q(apellidos__icontains=query) |
            Q(telefono__icontains=query)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    headers = ["Username", "Email", "Nombres", "Apellidos", "Teléfono", "Rol", "Estado"]
    ws.append(headers)

    for u in usuarios_qs.iterator():
        ws.append([
            u.username,
            u.email,
            u.nombres,
            u.apellidos,
            u.telefono or "",
            u.rol,
            u.estado
        ])

    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 2)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=usuarios.xlsx'
    wb.save(response)
    return response

# removed local generar_contrasena; using shared `generate_temporary_password`

@never_cache
@login_required
@role_required(['ADMIN'])
def crear_usuario(request): 
    if request.method == 'POST':
        form = UsuarioForm(request.POST)
        if form.is_valid():
            usuario = form.save(commit=False)
            # Generar contraseña aleatoria
            contrasena = generate_temporary_password()
            usuario.password = make_password(contrasena)
            # marcar que debe cambiar contraseña en el primer inicio
            usuario.must_change_password = True
            usuario.save()
            # Crear o sincronizar Django auth.User para manejo de sesiones
            if not AuthUser.objects.filter(username=usuario.username).exists():
                auth_user = AuthUser.objects.create_user(username=usuario.username, email=usuario.email)
                auth_user.set_password(contrasena)
                auth_user.save()
            else:
                auth_user = AuthUser.objects.get(username=usuario.username)
                auth_user.email = usuario.email
                auth_user.set_password(contrasena)
                auth_user.save()
            # Enviar correo con la contraseña temporal (para pruebas -> developer email)
            subject = "Cuenta creada - ERP Lili's"
            login_url = request.build_absolute_uri(reverse('login_registro'))
            message = (
                f"Se ha creado una cuenta para el usuario: {usuario.username}\n\n"
                f"Contraseña temporal: {contrasena}\n\n"
                f"Acceda aquí: {login_url}\n\n"
                "Al primer ingreso se le pedirá cambiar la contraseña."
            )
            from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', 'no-reply@example.com')
            # Send to the user's email if available; otherwise, don't attempt SMTP send.
            if usuario.email:
                try:
                    send_mail(subject, message, from_email, [usuario.email], fail_silently=getattr(settings, 'EMAIL_FAIL_SILENTLY', True))
                except smtplib.SMTPException as e:
                    logger = logging.getLogger(__name__)
                    logger.exception("Error enviando correo de creación de usuario: %s", e)
                    # Show a friendly message to the admin unless configured to be silent
                    if not getattr(settings, 'EMAIL_FAIL_SILENTLY', True):
                        messages.warning(request, 'Usuario creado, pero no se pudo enviar el correo (problema con SMTP). Revise la configuración de correo.')
            # Store the generated credentials in session and redirect to confirmation page.
            request.session['new_user_info'] = {'username': usuario.username, 'email': usuario.email, 'contrasena': contrasena}
            return redirect('usuario_confirmacion')
    else:
        form = UsuarioForm()
    return render(request, 'usuarios/crear_usuario.html', {'form': form})


@never_cache
@login_required
@role_required(['ADMIN'])
def admin_reset_password(request, pk):
    usuario = get_object_or_404(Usuario, pk=pk)
    if request.method != 'POST':
        messages.error(request, 'Operación inválida.');
        return redirect('lista_usuarios')

    # generar nueva clave temporal
    nueva = generate_temporary_password()
    usuario.password = make_password(nueva)
    usuario.must_change_password = True
    usuario.save()

    # sincronizar con auth.User si existe
    try:
        auth_user = AuthUser.objects.get(username=usuario.username)
        auth_user.set_password(nueva)
        auth_user.save()
    except AuthUser.DoesNotExist:
        pass

    # enviar correo de notificación (para pruebas -> developer email)
    subject = f"Reset de contraseña - cuenta {usuario.username}"
    login_url = request.build_absolute_uri(reverse('login_registro'))
    message = (
        f"Se ha generado una nueva contraseña temporal para el usuario: {usuario.username}\n\n"
        f"Contraseña temporal: {nueva}\n\n"
        f"Acceda aquí: {login_url}\n\n"
        "Al ingresar se solicitará cambiar la contraseña."
    )
    from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', 'no-reply@example.com')
    if usuario.email:
        try:
            send_mail(subject, message, from_email, [usuario.email], fail_silently=getattr(settings, 'EMAIL_FAIL_SILENTLY', True))
        except smtplib.SMTPException as e:
            logger = logging.getLogger(__name__)
            logger.exception("Error enviando correo de reset de contraseña: %s", e)
            if not getattr(settings, 'EMAIL_FAIL_SILENTLY', True):
                messages.warning(request, 'Se genero la contraseña pero no se pudo enviar el correo de notificación (problema SMTP).')

    messages.success(request, 'Se generó una nueva contraseña temporal y se envió un correo.')
    return redirect('lista_usuarios')


@never_cache
@login_required
@role_required(['ADMIN'])
def usuario_confirmacion(request):
    info = request.session.pop('new_user_info', None)
    if not info:
        messages.error(request, 'No hay información de usuario reciente para mostrar.')
        return redirect('lista_usuarios')
    usuario_obj = SimpleNamespace(username=info.get('username'), email=info.get('email'))
    return render(request, 'usuarios/confirmacion_usuario.html', {
        'usuario': usuario_obj,
        'contrasena': info.get('contrasena')
    })

def editar_usuario(request, pk):
    usuario = get_object_or_404(Usuario, pk=pk)
    if request.method == 'POST':
        form = UsuarioForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            return redirect('lista_usuarios')
    else:
        form = UsuarioForm(instance=usuario)
    return render(request, 'usuarios/editar_usuario.html', {'form': form, 'usuario': usuario})

#PRODUCTOS
@never_cache
@login_required
@role_required(['ADMIN'])
def crear_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_productos')
    else:
        form = ProductoForm()
    return render(request, 'productos/form.html', {'form': form, 'titulo': 'Agregar Producto'})

@never_cache
@login_required
@role_required(['ADMIN'])
def editar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            return redirect('lista_productos')
    else:
        form = ProductoForm(instance=producto)
    return render(request, 'productos/form.html', {'form': form, 'titulo': 'Editar Producto'})

@never_cache
@login_required
@role_required(['ADMIN'])
def eliminar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        producto.delete()
        return redirect('lista_productos')
    return render(request, 'productos/eliminar.html', {'producto': producto})


@never_cache
@login_required
@role_required(['ADMIN'])
def bulk_create_productos(request):
    """Create N test products for stress testing. POST only."""
    if request.method != 'POST':
        return redirect('lista_productos')

    try:
        count = int(request.POST.get('count', 1000))
    except (TypeError, ValueError):
        count = 1000
    count = max(1, min(count, 20000))

    objs = []
    # Avoid hitting validations/forms; create directly
    # Ensure we have a default category
    default_cat, _ = Categoria.objects.get_or_create(nombre_categoria='Sin categoría')

    for i in range(count):
        sku = f"TEST-{int(timezone.now().timestamp())}-{i}"
        nombre = f"Producto Prueba {i}"
        descripcion = "Carga masiva de prueba"
        precio = Decimal(str(round(random.uniform(100.0, 10000.0), 2)))
        stock = random.randint(0, 100)
        marca = "MarcaTest"
        objs.append(Producto(
            sku=sku,
            nombre=nombre,
            descripcion=descripcion,
            categoria=default_cat,
            marca=marca,
            modelo=None,
            uom_compra='UN',
            uom_venta='UN',
            factor_conversion=Decimal('1'),
            costo_estandar=None,
            costo_promedio=None,
            precio_venta=precio,
            impuesto_iva=Decimal('19'),
            stock=stock,
            stock_minimo=0,
            perishable=False,
            control_por_lote=False,
            control_por_serie=False,
        ))

    with transaction.atomic():
        Producto.objects.bulk_create(objs, batch_size=500)

    messages.success(request, f'Se crearon {count} productos de prueba.')
    return redirect('lista_productos')


@never_cache
@login_required
@role_required(['ADMIN'])
def bulk_delete_productos(request):
    """Delete products created by stress loader (heuristic: sku startswith 'TEST-')."""
    if request.method != 'POST':
        return redirect('lista_productos')

    qs = Producto.objects.filter(sku__startswith='TEST-')
    deleted_count, _ = qs.delete()
    messages.success(request, f'Se eliminaron {deleted_count} productos de prueba.')
    return redirect('lista_productos')
#VENTAS
@never_cache
@login_required
@role_required(['ADMIN','VENDEDOR','CAJA'])
def lista_ventas(request):
    query = request.GET.get('q', '').strip()
    ventas_qs = Venta.objects.all().order_by('-fecha')

    if query:
        ventas_qs = ventas_qs.filter(
            Q(id__icontains=query) |
            Q(fecha__icontains=query) |
            Q(total__icontains=query)
        )

    paginator = Paginator(ventas_qs, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {"page_obj": page_obj, "query": query}

    if request.headers.get("HX-Request") == "true":
        return render(request, "ventas/_grid_ventas.html", context)

    return render(request, "ventas/ventas.html", context)

@never_cache
@login_required
@role_required(['ADMIN','VENDEDOR','CAJA'])
def exportar_ventas_excel(request):

    query = request.GET.get('q', '').strip()
    ventas_qs = Venta.objects.all().order_by('-fecha')

    if query:
        ventas_qs = ventas_qs.filter(
            Q(id__icontains=query) |
            Q(fecha__icontains=query) |
            Q(total__icontains=query)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"

    headers = ["ID", "Fecha", "Total"]
    ws.append(headers)

    for v in ventas_qs.iterator():
        ws.append([v.id, v.fecha.strftime("%Y-%m-%d %H:%M"), v.total])

    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(h) + 2)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=ventas.xlsx'
    wb.save(response)
    return response

# NO SE USAN
def crear_venta(request):
    if request.method == 'POST':
        form = VentaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_ventas')
    else:
        form = VentaForm()
    return render(request, 'productos/form_venta.html', {'form': form, 'titulo': 'Agregar Venta'})

def editar_venta(request, pk):
    venta = get_object_or_404(Venta, pk=pk)
    if request.method == 'POST':
        form = VentaForm(request.POST, instance=venta)
        if form.is_valid():
            form.save()
            return redirect('lista_ventas')
    else:
        form = VentaForm(instance=venta)
    return render(request, 'productos/form_venta.html', {'form': form, 'titulo': 'Editar Venta'})

def eliminar_venta(request, pk):
    venta = get_object_or_404(Venta, pk=pk)
    if request.method == 'POST':
        venta.delete()
        return redirect('lista_ventas')
    return render(request, 'productos/eliminar_venta.html', {'venta': venta})
